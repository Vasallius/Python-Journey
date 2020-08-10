# Spreadsheet Cell Inverter

import openpyxl

# Load workbook
file_to_be_inverted = input("Enter file name to be inverted: ")
wb = openpyxl.load_workbook(file_to_be_inverted)

# Initialize variables
sheet = wb.active
cell_values = {}  # Dictionary in the form of {cellname:cellvalue}
cell_coord_list = []  # List containing all cells that had data before inverting

# Loop over the data in each column
for column_number in range(sheet.max_column):
    lst = list(sheet.columns)[column_number]
    for index, column in enumerate(lst):
        cell_coord = sheet.cell(row=column_number+1, column=index+1).coordinate
        # Add each cell coordinate to list
        cell_coord_list.append(column.coordinate)
        # Add cellname and its value to dictionary
        cell_values[cell_coord] = column.value

# Input values
for cellname in cell_coord_list:
    sheet[cellname] = cell_values.get(cellname, '')
for cellname in cell_values:
    sheet[cellname] = cell_values.get(cellname, '')

# Save workbook
wb.save('inverted.xlsx')
