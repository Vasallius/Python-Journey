# Spreadsheet to Textfiles

import openpyxl
from openpyxl.utils import get_column_letter

# Load workbook
excel_workbook = input("Enter name of workbook to convert to text: ")
wb = openpyxl.load_workbook(excel_workbook)
sheet = wb.active

for column_number in range(sheet.max_column):
    # Make new text file per column
    fh = open(f"{get_column_letter(column_number+1)}.txt", "a")

    # Iterate over each cell value in that column then write to the file
    for cell in list(sheet.columns)[column_number]:
        try:
            fh.write(cell.value)
        except:
            fh.write('')
    fh.close()
