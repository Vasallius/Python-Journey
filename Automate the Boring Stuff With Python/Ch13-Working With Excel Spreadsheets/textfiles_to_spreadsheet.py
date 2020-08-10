# Textfiles to Spreadsheet

import openpyxl
from openpyxl.utils import get_column_letter

# Make a new workbook
wb = openpyxl.Workbook()
sheet = wb.active

# Gather filenames
files = input("Enter filenames, separated by a space: ").split()
string_len_list = []

for x, filename in enumerate(files):
    stringlist = []
    fh = open(filename)

    # Append each string and its length to list
    for line in fh.readlines():
        stringlist.append(line)
        string_len_list.append(len(line))

    # Set the value of the cell to the string
    row_number = 1
    for string in stringlist:
        sheet[get_column_letter(x+1) + str(row_number)] = string
        sheet.column_dimensions[get_column_letter(
            x+1)].width = max(string_len_list)
        row_number += 1

wb.save('spreadsheet_from_text.xlsx')
