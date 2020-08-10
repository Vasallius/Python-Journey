# Multiplication Table Maker


import openpyxl
import sys
from openpyxl.styles import Font
from openpyxl.styles import Alignment

# Make a workbook
wb = openpyxl.Workbook(write_only=False)
sheet = wb.active

# Second argument to be used as number
number = int(sys.argv[1])
counter = 1

# Styling
bold_style = Font(bold=True)
center_style = Alignment(horizontal='center')

# Print the column headers (bold and centered)
for column_num in range(2, number+2):
    sheet.cell(row=1, column=column_num).value = counter
    sheet.cell(row=1, column=column_num).font = bold_style
    sheet.cell(row=1, column=column_num).alignment = center_style
    counter += 1

# Reset counter variable
counter = 1

# Print the row "headers" (bold and centered)
for row_num in range(2, number+2):
    sheet.cell(row=row_num, column=1).value = counter
    sheet.cell(row=row_num, column=1).font = bold_style
    sheet.cell(row=row_num, column=1).alignment = center_style
    counter += 1

# Print values of each cell
for row_num in range(2, number+2):
    for column_num in range(2, number+2):
        num1 = sheet.cell(row=row_num, column=1).value
        num2 = sheet.cell(row=1, column=column_num).value
        sheet.cell(row=row_num, column=column_num).value = num1 * num2
        sheet.cell(row=row_num, column=column_num).alignment = center_style

wb.save(f'{number}x{number} table.xlsx')
