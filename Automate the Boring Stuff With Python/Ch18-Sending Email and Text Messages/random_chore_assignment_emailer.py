# Random Chore Assignment Emailer

import smtplib
import openpyxl
import random

email_address = str(input('Email:'))
password = input('Password:')

wb = openpyxl.load_workbook('Chores.xlsx')
sheet = wb.active

names = []
chores = ['Clean the floor', 'Wash the dishes', 'Throw the trash']

for name in range(2, sheet.max_row+1):
    names.append(sheet.cell(row=name, column=1).value)
for chore in range(len(chores)):
    randomchore = random.choice(chores)
    # Make sure chore is not repeated
    while sheet.cell(row=chore+2, column=4).value == randomchore:
        randomchore = random.choice(chores)
    # Assign chores
    sheet.cell(row=chore+2, column=3).value = randomchore
    sheet.cell(row=chore+2, column=4).value = randomchore
    chores.remove(randomchore)

smtp_obj = smtplib.SMTP('smtp.gmail.com', 587)
smtp_obj.ehlo()
smtp_obj.starttls()
smtp_obj.login(user=email_address, password=password)

wb.save('Chores.xlsx')
# Send emails
for x in range(sheet.max_row-1):
    print(f'Sending mail to: {sheet.cell(row=x+2,column=2).value}')
    recipient_email = sheet.cell(row=x+2, column=2).value
    smtp_obj.sendmail(from_addr=email_address, to_addrs=recipient_email,
                      msg=f'Subject: {names[x]}\'s Chore of the Week\nPlease {sheet.cell(row=x+2,column=3).value}')

print('Emails have been sent.')
