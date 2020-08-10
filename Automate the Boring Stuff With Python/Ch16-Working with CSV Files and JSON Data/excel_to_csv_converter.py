# Excel to CSV Converter

import openpyxl
import csv
import os

for excel_file in os.listdir('.'):
    # Skip non-excel files
    if not excel_file.endswith('.xlsx'):
        continue

    # Load excel file
    wb = openpyxl.load_workbook(excel_file)

    # Loop through every sheet in the workbook.
    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]

        # Create the CSV filename from the Excel filename and sheet title.
        csvFile = open(f'{excel_file}_{sheetname}.csv', 'w', newline='')

        # Create the csv.writer object for this CSV file.
        outputWriter = csv.writer(csvFile)

        # Loop through every row in the sheet.
        for rowNum in range(1, sheet.max_row + 1):
            rowData = []

            # Loop through each cell in the row.
            for colNum in range(1, sheet.max_column + 1):
                # Append each cell's data to rowData.
                rowData.append(sheet.cell(row=rowNum, column=colNum).value)
            # Write the rowData list to the CSV file.
            outputWriter.writerow(rowData)
        csvFile.close()
